import aiosqlite
import db_config
import sqlite3
import openpyxl


async def create_table():
    async with aiosqlite.connect(db_config.get_database_path()) as db:
        cursor = await db.cursor()

        # Tworzenie tabeli active_systems, jeśli nie istnieje
        await cursor.execute('''
            CREATE TABLE IF NOT EXISTS active_systems (
                channel_id TEXT PRIMARY KEY,
                system_name TEXT,
                duration_hours INTEGER,
                clock_time TIMESTAMP,
                end_time TIMESTAMP
            )
        ''')

        # Tworzenie tabeli guild_timezones, jeśli nie istnieje
        await cursor.execute('''
            CREATE TABLE IF NOT EXISTS guild_config (
                guild_id TEXT PRIMARY KEY,
                timezone TEXT
            )
        ''')

        # Tworzenie tabeli guild_timezones, jeśli nie istnieje
        await cursor.execute('''
            CREATE TABLE IF NOT EXISTS time_zones (
                timezone TEXT PRIMARY KEY,
                utc_label TEXT
            )
        ''')

        # Tworzenie tabeli guild_timezones, jeśli nie istnieje
        await cursor.execute('''
            CREATE TABLE IF NOT EXISTS server_config (
                id INTEGER PRIMARY KEY,
                guild_id INTEGER UNIQUE,
                timezone TEXT,
                bot_nick TEXT,
                prefix TEXT DEFAULT "!",
                roll_prefix TEXT DEFAULT "k",
                welcome_channel INTEGER,
                log_channel INTEGER,
                moderation_role INTEGER,
                mute_role INTEGER,
                auto_role INTEGER,
                language TEXT DEFAULT "pl",
                localization TEXT DEFAULT "pl",
                roll_localization TEXT default "pl",
                removal_date TIMESTAMP
            )
        ''')

        # Tworzenie tabeli do sesji
        await cursor.execute('''
            CREATE TABLE IF NOT EXISTS game_sessions (
                session_id INTEGER PRIMARY KEY AUTOINCREMENT,
                channel_id TEXT NOT NULL,
                system_name TEXT NOT NULL,
                gm_id TEXT NOT NULL,
                gm_nick TEXT NOT NULL,
                player_id TEXT,
                player_nick TEXT,
                active BOOLEAN DEFAULT TRUE
            )
        ''')

        # Tworzenie tabeli dla graczy
        await cursor.execute('''
            CREATE TABLE IF NOT EXISTS session_players (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id INTEGER,
                player_id TEXT NOT NULL,
                player_nick TEXT NOT NULL,
                FOREIGN KEY(session_id) REFERENCES game_sessions(session_id)
            )
        ''')

        await db.commit()


# Zaczytywanie UTC
def load_data_from_excel_to_db(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Zliczanie liczby wierszy w arkuszu (bez nagłówka)
    excel_row_count = sum(1 for row in sheet.iter_rows(values_only=True)) - 1

    conn = sqlite3.connect(db_config.get_database_path())
    cursor = conn.cursor()

    # Zliczanie liczby wpisów w tabeli time_zones
    cursor.execute("SELECT COUNT(*) FROM time_zones")
    db_row_count = cursor.fetchone()[0]

    # Jeśli liczba wierszy w arkuszu i tabeli jest różna, wczytaj dane
    if excel_row_count != db_row_count:
        print(f"Aktualizuję bazę UTC")
        cursor.execute("DELETE FROM time_zones")  # Czyszczenie tabeli
        for row in sheet.iter_rows(min_row=2, values_only=True):  # pomija nagłówek
            city = row[0]
            utc_label = row[1]
            cursor.execute("INSERT INTO time_zones (timezone, utc_label) VALUES (?, ?)", (city, utc_label))
        conn.commit()

    conn.close()


load_data_from_excel_to_db('UTC.xlsx')
